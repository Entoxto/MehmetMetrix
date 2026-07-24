"""Регрессионные тесты критической логики Excel-парсера."""

import unittest
from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import Workbook

from excel_parser import ExcelParser
from parser_utils import parse_sizes_from_name


class ParserLogicTests(unittest.TestCase):
    def create_row(self, exchange_rate=None, cost=42000):
        row = pd.Series([None] * 16)
        row.iloc[2] = "Жакет из кожи — тестовый (XS-1)"
        row.iloc[7] = 100
        row.iloc[9] = exchange_rate
        row.iloc[13] = cost
        return row

    def parse_item(self, row):
        with redirect_stdout(StringIO()):
            return ExcelParser("unused.xlsx", [])._parse_item(row, excel_row=2)

    def test_duplicate_size_stops_parsing_with_excel_row(self):
        with self.assertRaisesRegex(
            ValueError,
            r"Строка 57: обнаружено дублирование размера XS",
        ):
            parse_sizes_from_name("Жакет из кожи (XS-2, XS-3)", excel_row=57)

    def test_duplicate_one_size_stops_parsing_with_excel_row(self):
        with self.assertRaisesRegex(
            ValueError,
            r"Строка 58: обнаружено дублирование размера ONE SIZE",
        ):
            parse_sizes_from_name(
                "Жакет из кожи (one size-2, OneSize-3)",
                excel_row=58,
            )

    def test_blank_exchange_rate_does_not_import_cost(self):
        item = self.parse_item(self.create_row())

        self.assertNotIn("cost", item)

    def test_non_positive_exchange_rate_does_not_import_cost(self):
        for exchange_rate in (0, -1):
            with self.subTest(exchange_rate=exchange_rate):
                item = self.parse_item(self.create_row(exchange_rate=exchange_rate))
                self.assertNotIn("cost", item)

    def test_positive_exchange_rate_imports_cost(self):
        item = self.parse_item(self.create_row(exchange_rate=95.5))

        self.assertEqual(item["cost"], 42000)

    def test_merged_exchange_rate_imports_cost_for_every_position(self):
        with TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "shipments.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Поставки"

            for column in range(1, 15):
                worksheet.cell(row=1, column=column, value=f"column-{column}")

            worksheet.cell(row=2, column=1, value=1)
            worksheet.cell(
                row=2,
                column=3,
                value="Жакет из кожи — первый (XS-1)",
            )
            worksheet.cell(row=2, column=7, value=1)
            worksheet.cell(row=2, column=8, value=100)
            worksheet.cell(row=2, column=10, value=95.5)
            worksheet.cell(row=2, column=14, value=42000)

            worksheet.cell(
                row=3,
                column=3,
                value="Жакет из кожи — второй (XS-1)",
            )
            worksheet.cell(row=3, column=7, value=1)
            worksheet.cell(row=3, column=8, value=100)
            worksheet.cell(row=3, column=14, value=43000)
            worksheet.merge_cells("J2:J3")

            workbook.save(excel_path)
            workbook.close()

            with redirect_stdout(StringIO()):
                shipments = ExcelParser(str(excel_path), []).parse()

        costs = [
            item.get("cost")
            for item in shipments[0]["rawItems"]
        ]
        self.assertEqual(costs, [42000, 43000])


if __name__ == "__main__":
    unittest.main()
